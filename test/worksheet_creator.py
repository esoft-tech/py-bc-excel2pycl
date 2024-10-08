from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date, timedelta


def create_test_table(file_name):
    wb = Workbook()
    ws = wb.active

    data = [
        # ROUND - 2 строка
        ['=ROUND(A4, 2)', '=OR(B4, B5)', '=AND(C4, C5)', '=VLOOKUP(5, D4:D13, 1, FALSE())', '=IF(E6>E5, 44, 11)',
         '=SUM(F4:F13)', '=SUMIF(G4:G13, ">6")', '=AVERAGE(H4:H13)',
         '=MIN(I4:I13)', '=MAX(J4:J13)',
         '=YEAR(K4)', '=MONTH(L4)', '=DAY(M4)',
         '=IFERROR(N4/N5,5)', '=IFERROR(O4+O5,5)',
         '=EOMONTH(P4, 2)', '=EOMONTH(Q4, -2)',
         '=EDATE(R4, 2)', '=EDATE(S4, -1)', '=MATCH(30;T4:T12;1)', '=XMATCH(50;U4:U12;1;1)'],
        [10.24, '=TRUE()', '=FALSE()', 5, 44, 55, 34, 5.5, 1, 10, 2023, 10, 10, 5, 12, datetime(
            2022, 3, 31), datetime(2021, 11, 30), datetime(2011, 3, 15), datetime(2010, 12, 15), 4, 6],
        [10.239584, '', 0, 1, 1, 1, 1, 1, 1, 1, datetime(2023, 10, 10), datetime(2023, 10, 10), datetime(
            2023, 10, 10), 12, 12, datetime(2022, 1, 1), datetime(2022, 1, 1), datetime(2011, 1, 15), datetime(2011, 1, 15), 0, 0],
        ['', 100, 1, 2, 2, 2, 2, 2, 2, 2, '', '', '', 0, 0, '', '', '', '', 10, 10],
        ['', '', '', 3, 3, 3, 3, 3, 3, 3, '', '', '', '', '', '', '', '', '', 20, 20],
        ['', '', '', 4, 4, 4, 4, 4, 4, 4, '', '', '', '', '', '', '', '', '', 30, 30],
        ['', '', '', 5, 5, 5, 5, 5, 5, 5, '', '', '', '', '', '', '', '', '', 40, 40],
        ['', '', '', 6, 6, 6, 6, 6, 6, 6, '', '', '', '', '', '', '', '', '', 50, 50],
        ['', '', '', 7, 7, 7, 7, 7, 7, 7, '', '', '', '', '', '', '', '', '', 60, 60],
        ['', '', '', 8, 8, 8, 8, 8, 8, 8, '', '', '', '', '', '', '', '', '', 70, 70],
        ['', '', '', 9, 9, 9, 9, 9, 9, 9, '', '', '', '', '', '', '', '', '', 80, 80],
        ['', '', '', 10, 10, 10, 10, 10, 10, 10, '', '', '', '', '', '', '', '', '', 90, 90],
    ]

    # add column headings. NB. these must be strings
    ws.append(["ROUND", "OR", "AND", "VLOOKUP", "IF", "SUM", "SUMIF", "AVERAGE", "MIN",
               "MAX", "YEAR", "MONTH", "DAY",
               "IFERROR when_error", "IFERROR condition",
               "EOMONTH", "EOMONTH NEGATIVE",
               "EDATE plus", "EDATE minus", "MATCH", "XMATCH"])

    for row in data:
        ws.append(row)

    tab = Table(displayName='base', ref='A1:E5')

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    ws.add_table(tab)

    ws_date = wb.create_sheet('date')

    data = [
        ['=DATE(A4, A5, A6)', '=DATE(B4, B5, B6)', '=DATE(C4, C5, C6)', '=DATE(D4, D5, D6)',
         '=DATE(E4, E5, E6)', '=DATE(F4, F5, F6)', '=DATE(G4, G5, G6)', '=DATE(H4, H5, H6)', '=DATE(I4, I5, I6)',
         '=DATE(J4, J5, J6)'],
        [datetime(2022, 10, 10), datetime(2000, 10, 10), '#NUM!', datetime(2023, 1, 1), datetime(2024, 6, 1),
         datetime(2021, 11, 1), datetime(2019, 6, 1), datetime(2022, 8, 28), datetime(2022, 3, 17),
         datetime(2022, 9, 29)],
        [2022, 100, 10000, 2022, 2022, 2022, 2022, 2022, 2022, 2022],
        [10, 10, 10, 13, 30, -1, -30, 5, 5, 10],
        [10, 10, 10, 1, 1, 1, 1, 120, -44, -1],
    ]

    # add column headings. NB. these must be strings
    ws_date.append(["DATE normal", "DATE year < 1900", "DATE year > 9999",
                    "DATE month > 12", "DATE month > 24", "DATE month < 1", "DATE month < -12",
                    "DATE days > 30", "DATE days < 0", "DATE days = -1"])
    for row in data:
        ws_date.append(row)

    ws_datedif = wb.create_sheet('datedif')

    data = [
        ['=DATEDIF(A4, A5, "D")', '=DATEDIF(B4, B5, "M")', '=DATEDIF(C4, C5, "Y")', '=DATEDIF(D4, D5, "MD")',
         '=DATEDIF(E4, E5, "YM")', '=DATEDIF(F4, F5, "YD")', '=DATEDIF(G4, G5, "YD")', '=DATEDIF(H4, H5, "YD")'],
        [145, 4, 0, 30, 6, 145, 14, 337],
        [datetime(2022, 10, 10), datetime(2022, 10, 10), datetime(2022, 10, 10),
         datetime(2022, 10, 10), datetime(2022, 3, 14), datetime(2022, 10, 10), datetime(2022, 2, 27),
         datetime(2022, 4, 10)],
        [datetime(2023, 3, 4), datetime(2023, 3, 4), datetime(2023, 3, 4),
         datetime(2023, 2, 9), datetime(2023, 10, 5), datetime(2023, 3, 4), datetime(2023, 3, 13),
         datetime(2023, 3, 13)],
    ]

    # add column headings. NB. these must be strings
    ws_datedif.append(["DATEDIF D", "DATEDIF M", "DATEDIF Y", "DATEDIF MD", "DATEDIF YM", "DATEDIF YD", "DATEDIF YD 2",
                       "DATEDIF YD 3"])
    for row in data:
        ws_datedif.append(row)

    ws_left = wb.create_sheet('left')

    data = [
        ['LEFT some num', 'LEFT zero', 'LEFT big length', 'LEFT eq len', 'LEFT one arg', 'LEFT negative arg'],
        ['=LEFT(A4,2)', '=LEFT(B4,0)', '=LEFT(C4,200)', '=LEFT(D4,11)', '=LEFT(E4)', '=LEFT(F4,-2)' ],
        ['He', '', 'Hello World', 'Hello World', 'H', '#ERROR!'],
        ['Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World']
    ]
    
    for row in data:
        ws_left.append(row)
        
        
    ws_right = wb.create_sheet('right')

    data = [
        ['RIGHT some num', 'RIGHT zero', 'RIGHT big length', 'RIGHT eq len', 'RIGHT one arg', 'RIGHT negative arg'],
        ['=RIGHT(A4,5)', '=RIGHT(B4,0)', '=RIGHT(C4,200)', '=RIGHT(D4,11)', '=RIGHT(E4)', '=RIGHT(F4,-2)' ],
        ['World', '', 'Hello World', 'Hello World', 'd', '#ERROR!'],
        ['Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World']
    ]
    
    for row in data:
        ws_right.append(row)
        
        
    ws_mid = wb.create_sheet('mid')

    data = [
        ['MID normal', 'MID zero', 'MID big length', 'MID full length', 'MID eq length', 'MID sec arg err', 'MID third arg err'],
        ['=MID(A4, 1, 5)', '=MID(B4, 1, 0)', '=MID(C4, 100, 100)', '=MID(D4, 1, 100)', '=MID(E4, 1, 11)', '=MID(F4,0,10)', '=MID(F4,1,-1)' ],
        ['Hello', '', '', 'Hello World', 'Hello World', '#NUM!', '#VALUE!'],
        ['Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World']
    ]
    
    for row in data:
        ws_mid.append(row)

    ws_mid = wb.create_sheet('count')

    data = [
        ['COUNT normal', 'COUNT single cell', 'COUNT num & string digits', 'COUNT range & arg sequence',
         'COUNT range & arg sequence with bool & string digits & single cells'],
        ['=COUNT(B3:I3)', '=COUNT(B4; C4)', '=COUNT(B5:H5)', '=COUNT(B6:H6; 2; 3)',
         '=COUNT(B7:H7; TRUE; FALSE; "asd"; "2"; I7; J7)', '=COUNT(B8:H8)'],
        [2, 'Hello', 1, 2, 'Hello World', 'Hello World', '#NUM!', '#VALUE!', '#DIV/0'],
        [1, 1, 'Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World', 'Hello World'],
        [2, 1, 2, 'Hello World', '4', '6', 'Hello World', 'Hello World'],
        [3, 1, '2', 'Hello World', '4', '6', 'Hello World', 'Hello World'],
        [5, 1, '2', 'Hello World', '4', '6', 'Hello World', '0', 2, '23'],
        [2, 1, '2', datetime(2022, 10, 10), '4', '6', 'Hello World', '0'],
    ]

    for row in data:
        ws_mid.append(row)

    ws_countifs = wb.create_sheet('countifs')

    data = [
        ['COUNTIFS text condition', 'COUNTIFS cell condition', 'COUNTIFS lambda condition',
         'COUNTIFS expression condition', 'COUNTIFS any text condition', 'COUNTIF pattern text condition'],
        ['=COUNTIFS(A4:A7;"яблоки";B4:B7;">55")', '=COUNTIFS(A4:A7;A5)', '=COUNTIFS(B4:B7;">55")',
         '=COUNTIFS(B4:B7;"<>"&B5)', '=COUNTIFS(A4:A7;"*")', '=COUNTIFS(A4:A7;"????ки")'],
        [1, 1, 2, 3, 4, 2],
        ['яблоки', 32],
        ['апельсины', 54],
        ['персики', 75],
        ['яблоки', 86],
    ]

    for row in data:
        ws_countifs.append(row)

    ws_address = wb.create_sheet('address')

    data = [
        ['ADDRESS absolute', 'ADDRESS absolute row, relative col', 'ADDRESS absolute col, relative row',
         'ADDRESS relative', 'ADDRESS absolute strict', 'ADDRESS RC type col relative',
         'ADDRESS RC type col relative link sheet', 'ADDRESS absolute link sheet & workbook', 'ADDRESS HUGE'],
        ['=ADDRESS(3;1)', '=ADDRESS(3;2;2)', '=ADDRESS(3;3;3)', '=ADDRESS(3;4;4)', '=ADDRESS(3;5;1)',
         '=ADDRESS(3;6;2;FALSE)', '=ADDRESS(3;7;2;FALSE;"mid")', '=ADDRESS(3;8;1;TRUE;"[WorkBook1]ASD")',
         '=ADDRESS(3;704)'],
        ['$A$3', 'B$3', '$C3', 'D3', '$E$3', 'R3C[6]', "'mid'!R3C[7]", "'[WorkBook1]ASD'!$H$3", '$AAB$3'],
    ]

    for row in data:
        ws_address.append(row)

    ws_today = wb.create_sheet('today')
    data = [
        ['TODAY normal', 'TODAY ADD DAY', 'TODAY subtract today from date', 'Get DAY from TODAY',
         'Get MONTH from TODAY', 'Compare TODAY and TODAY'],
        ['=TODAY()', '=TODAY() + 5', '=DATE(2024, 5, 24) - TODAY()', '=DAY(TODAY())',
         '=MONTH(TODAY())', '=TODAY() = TODAY()'],
        [
            date.today(),
            date.today() + timedelta(days=5),
            (date(2024, 5, 24) - date.today()).days,
            date.today().day,
            date.today().month,
            date.today() == date.today()
        ]
    ]

    for row in data:
        ws_today.append(row)

    ws_column = wb.create_sheet('column')

    data = [
        ['COLUMN no args'],
        [1, 2, 3, 4, 5, 3],
        ['=COLUMN()'],
        ['=COLUMN(B3)'],
        ['=COLUMN(C3:E3)'],
        ['=COLUMN(C3:C7)']
    ]

    for row in data:
        ws_column.append(row)

    ws_index = wb.create_sheet('index')

    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
        ['=INDEX(A1:C1;2)', '=INDEX(A1:A3;2)', '=INDEX(A1:C3;2;2)', '=INDEX((A1:C1; A1:A3; A1:C3);3;3;3)'],
        ['=INDEX(A1:C3;0;0)'],
        ['=INDEX(A1:A3&B1:B3, 0)']
    ]

    for row in data:
        ws_index.append(row)

    ws_mid = wb.create_sheet('sumifs')

    data = [
        ['Quantity Sold', 'Product', 'Salesperson', '=SUMIFS($A$2:$A$9, $C$2:$C$9, "Tom")', 52],
        [5, 'Apples', 'Tom', '=SUMIFS($A$2:$A$9, $C$2:$C$9, "Tom", B2:B9, "Bananas")', 22],
        [3, 'Apples', 'Sarah', '=SUMIFS($A$2:$A$9, $B$2:$B$9, "A*")', 26],
        [15, 'Artichokes', 'Tom', '=SUMIFS($A$2:$A$11, $C$2:$C$11, "Tom", B2:B11, "Bananas")', 23],
        [3, 'Artichokes', 'Sarah', '=SUMIFS($A$2:$A$9, $A$13:$A$20, "Tom")', 0],
        [22, 'Bananas', 'Tom', '=SUMIFS($A$2:$A$11, C2:C11, "Sarah", C2:C11, "Tom")', 0],
        [12, 'Bananas', 'Sarah', '=SUMIFS($A$2:$A$9, $A$2:$A$9, ">10")', 82],
        [10, 'Carrots', 'Tom', '=SUMIFS($A$2:$A$9, $A$2:$A$9, "<10")', 11],
        [33, 'Carrots', 'Sarah'],
        [True, 'Bananas', 'Tom'],
        [False, 'Artichokes', 'Tom'],
    ]

    for row in data:
        ws_mid.append(row)

    ws_roundup = wb.create_sheet('roundup')

    data = [
        ['=ROUNDUP(C1)', '=ROUNDUP(C1, 1)', 3.14],
        ['=ROUNDUP(C2)', '=ROUNDUP(C2, 2)', 3.1415926],
        ['=ROUNDUP(C3)', '=ROUNDUP(C3, 2)', 3.141],
        ['=ROUNDUP(C1,)'],
        ['=ROUNDUP(C2,)'],
        ['=ROUNDUP(C3,)']
    ]

    for row in data:
        ws_roundup.append(row)

    ws_rounddown = wb.create_sheet('rounddown')

    data = [
        ['=ROUNDDOWN(C1)', '=ROUNDDOWN(C1, 1)', 3.14],
        ['=ROUNDDOWN(C2)', '=ROUNDDOWN(C2, 2)', 3.1415926],
        ['=ROUNDDOWN(C3)', '=ROUNDDOWN(C3, 2)', 3.141],
        ['=ROUNDDOWN(C1,)'],
        ['=ROUNDDOWN(C2,)'],
        ['=ROUNDDOWN(C3,)']
    ]

    for row in data:
        ws_rounddown.append(row)

    ws_left_operand_expression = wb.create_sheet('leftOperandExpression')

    data = [
        ['=B1%', 7],
        ['=B1%*2', 12],
        ['=B1%*B2', '=7%'],
        ['=B1%*B2%'],
        ['=B3%']
    ]

    for row in data:
        ws_left_operand_expression.append(row)

    ws_value = wb.create_sheet('value')

    data = [
        ['=VALUE("123")', '2024-09-15'],
        ['=VALUE("123.45")'],
        ['=VALUE("2024-09-15")'],
        ['=VALUE("12:30:00")'],
        ['=VALUE(50% /100)'],
        ['=VALUE("123abc")'],
        ['=VALUE(B1)']
    ]

    for row in data:
        ws_value.append(row)

    ws_text = wb.create_sheet('text')
    data = [
        ['=TEXT("1234567", "#,##0")'],
    ]

    for row in data:
        ws_text.append(row)

    ws_concatenate = wb.create_sheet('concatenate')
    data = [
        ['=CONCATENATE("при", B1)', 'вет'],
        ['=CONCATENATE("при", 123, "вет")'],
        ['=CONCATENATE("при", DATE(2024, 9, 12), "вет")']
    ]

    for row in data:
        ws_concatenate.append(row)

    wb.save(file_name)
