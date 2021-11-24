from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from src.cell import Cell


class Excel:
    def __init__(self, worksheets):
        self._data = worksheets['data']
        self._titles = {title: worksheet_number for title, worksheet_number
                        in zip(worksheets['titles'], range(len(worksheets['titles'])))}

    def _title_to_number(self, title: str) -> int:
        return self._titles[title]

    def handle_cell(self, cell: Cell):
        if type(cell.title) is str:
            cell.title = self._title_to_number(cell.title)

        if type(cell.column) is str:
            cell.column = column_index_from_string(cell.column) - 1

    def _get_cell(self, title: int, column: int, row: int) -> int or float or str or bool or None:
        return self._data[title][row][column] if 0 <= title < len(
            self._data) and 0 <= row < len(self._data[title]) and 0 <= column < len(
            self._data[title][row]) else None

    def get_cell(self, cell: Cell) -> int or float or str or bool or None:
        if cell.row is None:
            # TODO добавить кастомные исключения
            raise Exception('It is not possible to get a cell without pointing to a specific row')

        self.handle_cell(cell)

        return self._get_cell(cell.title, cell.column, cell.row)

    # TODO should we return set if the function name is get_set?
    # TODO добавить проверки на предмет выхода за диапазоны excel-файлика
    # TODO добавить проерки на предмет того, что дальше, а что ближе
    def get_set(self, first: Cell, second: Cell) -> list:
        self.handle_cell(first)
        self.handle_cell(second)

        if first.title != second.title:
            raise Exception('It is impossible to get set if the values are located in different workspaces')

        if first.column == second.column:
            result = self._get_vertical_set(first, second)
        elif first.row == second.row:
            result = self._get_horizontal_set(first, second)
        else:
            raise Exception('It is impossible to get a set if its values are not located in a straight line')

        return result

    def _get_vertical_set(self, first: Cell, second: Cell) -> list:
        start_row = first.row
        finish_row = second.row
        if start_row is None:
            start_row = 0
            finish_row = len(self._data[first.title])
        else:
            finish_row += 1

        result = []
        for row in range(start_row, finish_row):
            result.append(self._get_cell(first.title, first.column, row))

        return result

    def _get_horizontal_set(self, first: Cell, second: Cell) -> list:
        start_column = first.column
        finish_column = second.column + 1

        result = []
        for column in range(start_column, finish_column):
            result.append(self._get_cell(first.title, column, first.row))

        return result

    # TODO добавить проверки аналогичные get_set
    def get_matrix(self, first: Cell, second: Cell) -> list:
        self.handle_cell(first)
        self.handle_cell(second)

        if first.title != second.title:
            raise Exception('It is impossible to get matrix if the values are located in different workspaces')

        result = []
        for row in range(first.row, second.row + 1):
            row_data = []
            for column in range(first.column, second.column + 1):
                row_data.append(self._get_cell(first.title, column, row))
            result.append(row_data)

        return result

    @classmethod
    def parse(cls, path: str):
        worksheets_data = []
        worksheets_titles = []
        wb = load_workbook(filename=path)
        for worksheet in wb.worksheets:
            worksheets_titles.append(worksheet.title)
            worksheet_data = []
            last_column = len(list(worksheet.columns))
            last_row = len(list(worksheet.rows))
            for row_number in range(1, last_row + 1):
                rows_data = []
                for column_number in range(1, last_column + 1):
                    column_letter = get_column_letter(column_number)
                    rows_data.append(worksheet[f'{column_letter}{row_number}'].value)
                worksheet_data.append(rows_data)
            worksheets_data.append(worksheet_data)

        return cls({'data': worksheets_data, 'titles': worksheets_titles})
