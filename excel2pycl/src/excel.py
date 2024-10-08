import re
from typing import List, Dict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from excel2pycl.src.cell import Cell
from excel2pycl.src.exceptions import E2PyclSafetyException, E2PyclParserException
from excel2pycl.src.handle_cell import handle_cell


class Excel:
    def __init__(self, worksheets):
        self._data = worksheets['data']
        self._titles = {title: worksheet_number for title, worksheet_number
                        in zip(worksheets['titles'], range(len(worksheets['titles'])))}
        self._suspicious_cells = worksheets['suspicious_cells']
        self._sheets_size = worksheets['sheets_size']

    def is_safe(self):
        """
        Throws as exception if Excel file contains Python-like content
        """
        if self._suspicious_cells:
            raise E2PyclSafetyException(suspicious_cells=self._suspicious_cells)

    @staticmethod
    def _handle_cell(cell: Cell):
        """
        If cell hasn't only integer identifiers, throws exception
        """
        cell.uid

    def _fill_cell(self, cell: Cell) -> Cell:
        self._handle_cell(cell)
        cell.value = self._data[cell.title][cell.row][cell.column] if 0 <= cell.title < len(
            self._data) and 0 <= cell.row < len(self._data[cell.title]) and 0 <= cell.column < len(
            self._data[cell.title][cell.row]) else None
        return cell

    def fill_cell(self, cell: Cell) -> Cell:
        if cell.row is None:
            # TODO добавить кастомные исключения
            raise E2PyclParserException('It is not possible to get a cell without pointing to a specific row')

        handle_cell(cell, self._titles)

        return self._fill_cell(cell)

    # TODO should we return set if the function name is get_set?
    # TODO добавить проверки на предмет выхода за диапазоны excel-файлика
    # TODO добавить проверки на предмет того, что дальше, а что ближе
    def get_range(self, first: Cell, second: Cell) -> list:
        handle_cell(first, self._titles)
        handle_cell(second, self._titles)

        if first.title != second.title:
            raise E2PyclParserException(
                'It is impossible to get range if the values are located in different workspaces')

        if first.column == second.column:
            result = self._get_vertical_range(first, second)
        elif first.row == second.row:
            result = self._get_horizontal_range(first, second)
        else:
            raise E2PyclParserException(
                'It is impossible to get a range if its values are not located in a straight line')

        return result

    def get_similar_second(self, base: Cell, first: Cell, second: Cell):
        handle_cell(base, self._titles)
        handle_cell(first, self._titles)
        handle_cell(second, self._titles)

        return Cell(base.title, base.column + (second.column - first.column), base.row + (second.row - first.row) if first.row is not None or second.row is not None else None)

    def _get_vertical_range(self, first: Cell, second: Cell) -> list:
        start_row = first.row
        finish_row = second.row
        if start_row is None:
            start_row = 0
            finish_row = len(self._data[first.title])
        else:
            finish_row += 1

        result = []
        for row in range(start_row, finish_row):
            result.append(self._fill_cell(Cell(title=first.title, column=first.column, row=row)))

        return result

    def _get_horizontal_range(self, first: Cell, second: Cell) -> list:
        start_column = first.column
        finish_column = second.column + 1

        result = []
        for column in range(start_column, finish_column):
            result.append(self._fill_cell(Cell(title=first.title, column=column, row=first.row)))

        return result

    def _get_matrix(self, first: Cell, second: Cell) -> list:
        if first.title != second.title:
            raise E2PyclParserException(
                'It is impossible to get matrix if the values are located in different workspaces')

        result = []
        for row in range(first.row, second.row + 1):
            row_data = []
            for column in range(first.column, second.column + 1):
                row_data.append(self._fill_cell(Cell(title=first.title, column=column, row=row)))
            result.append(row_data)

        return result

    # TODO добавить проверки аналогичные get_set
    def get_matrix(self, first: Cell, second: Cell) -> list:
        handle_cell(first, self._titles)
        handle_cell(second, self._titles)

        if first.row is None and second.row is None:
            if first.column == second.column:
                # A:A range case
                return [[i] for i in self._get_vertical_range(first, second)]
            # A:C range case
            result = list((self._get_vertical_range(Cell(first.title, column_index, None), Cell(
                first.title, column_index, None)) for column_index in range(first.column, second.column+1)))
            return result
        elif isinstance(first.row, int) and first.row >= 0 and second.row >= 0:
            return self._get_matrix(first, second)
        else:
            raise E2PyclParserException('Invalid cell coordinates')

    @classmethod
    def _get_suspicious_constructions(cls, value):
        value = str(value)
        suspicious_constructions = re.findall(r'[a-zA-Z_\d]+\(.*?\)', value)
        if suspicious_constructions:
            return [i for i in suspicious_constructions if not re.findall(r'[A-Z]+\(.*?\)', i)]

        return []

    @classmethod
    def parse(cls, path: str):
        worksheets_data = []
        worksheets_titles = []
        suspicious_cells = {}
        sheets_size = []
        wb = load_workbook(filename=path, read_only=True)
        for worksheet in wb.worksheets:
            worksheet.reset_dimensions()
            worksheets_titles.append(worksheet.title)
            worksheet_data = []
            max_row_len = 0
            for row in worksheet.iter_rows():
                rows_data = []
                for index, cell in enumerate(row):
                    if cell.value and (suspicious_constructions := cls._get_suspicious_constructions(cell.value)):
                        suspicious_cells[f"'{worksheet.title}'{cell.column_letter}{index+1}"] = suspicious_constructions

                    # обрабатываем ArrayFormula, считываем из него значение формулы
                    if isinstance(cell.value, ArrayFormula):
                        rows_data.append(cell.value.text.strip())
                    else:
                        rows_data.append(cell.value)
                worksheet_data.append(rows_data)
                rows_data_len = len(rows_data)
                if max_row_len < rows_data_len:
                    max_row_len = rows_data_len
            sheets_size.append({'last_column': max_row_len, 'last_row': len(worksheet_data)})
            worksheets_data.append(worksheet_data)
        wb.close()

        return cls({
            'data': worksheets_data,
            'titles': wb.sheetnames,
            'suspicious_cells': suspicious_cells,
            'sheets_size': sheets_size,
        })

    def get_cells(self) -> List[Cell]:
        """
        Returns all cells from Excel file.

        Returns:
            List[Cell]: List of filled cells.
        """
        cells: List[Cell] = []
        for title_number, title in enumerate(self._data):
            for row_number, row in enumerate(title):
                for column_number, column in enumerate(row):
                    cells.append(self.fill_cell(Cell(title_number, column_number, row_number)))

        return cells

    def get_titles(self) -> Dict[str, int]:
        return self._titles

    def get_sheets_size(self) -> list[Dict[str, int]]:
        return self._sheets_size
